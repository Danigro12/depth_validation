import pandas as pd
import numpy as np
import openpyxl
import json
from Bio import SeqIO
import sys
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule

#Variables
path_list = sys.argv[1]
report_name = sys.argv[2] if len(sys.argv) > 2 else "mosdepth_validation"
path_to_file = []
tabela_final = {
    'classification': pd.DataFrame(),
    'scores': pd.DataFrame()
}

#Parsing input files
with open(path_list,'r') as input:
    parse_lines = input.readlines()
    for line in parse_lines:
        path_to_file.append(line.strip())

def importar_dataset(path_to_file):
    file_name = path_to_file.strip().split('/')[-1]
    file_name = file_name.replace(".thresholds.bed.gz", "")
    stats = {}

    #Importing Mosdepth Output and selecting its columns
    mosdepth_output = pd.read_table(path_to_file, sep='\t', compression = 'gzip')
    mosdepth_output = mosdepth_output.loc[:, ['#chrom', 'start', 'end', 'region', '0X', '30X']]

    #Extract Gene Name by regex
    mosdepth_output['gene_name'] = mosdepth_output['region'].str.extract(r"^([^.]*)")

    #Setting validation column, to see x30 depth of which exon
    mosdepth_output[f'score_{file_name}'] = mosdepth_output['30X'] - mosdepth_output['0X']

    gene_size = mosdepth_output.groupby('region', as_index=False)['0X'].agg('sum')

    #Aggregation of values from which exon, to get depth by gene
    df_agrupado = mosdepth_output.groupby('region', as_index=False)[f'score_{file_name}'].agg('sum')
    df_agrupado[f'{file_name}'] = df_agrupado[f'score_{file_name}'].apply(
        lambda x: 'OK' if x == 0 else ('OUT' if x < 0 else np.nan))  # Changed NA to np.nan

    df_agrupado[f'score_{file_name}'] = (df_agrupado[f'score_{file_name}'].astype(str) + "/" + gene_size['0X'].astype(str))

    # Building the dict with dataframes
    if len(tabela_final['classification']) == 0:
        tabela_final['classification'] = df_agrupado[['region', f'{file_name}']] 
    else:
        tabela_final['classification'] = pd.merge(
            tabela_final['classification'],
            df_agrupado[['region', f'{file_name}']],  
            on="region",
            how="left"
        )
    
    if len(tabela_final['scores']) == 0:
        tabela_final['scores'] = df_agrupado[['region', f'score_{file_name}']]  
    else:
        tabela_final['scores'] = pd.merge(
            tabela_final['scores'],
            df_agrupado[['region', f'score_{file_name}']],  
            on="region",
            how="left"
        )

for file in path_to_file:
    importar_dataset(file)

with pd.ExcelWriter(f'md_report_{report_name}.xlsx', engine = 'openpyxl') as writer:
    for nome_aba, df in tabela_final.items():
        df.to_excel(writer,sheet_name=nome_aba,index=False)

wb = openpyxl.load_workbook(f'md_report_{report_name}.xlsx')
for sheet_name in wb.sheetnames:
    if sheet_name.startswith("classification"):
        ws = wb[sheet_name]
        for col in ws.iter_cols(min_row=2, min_col=2):
            for cell in col:
                if cell.value == "OUT":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif cell.value == "OK":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
wb.save(f'md_report_{report_name}.xlsx')